import pandas as pd
import numpy as np
import os
import json
import time
from data_processing import filter_dataframes
from survey_error import survey_se
from openpyxl.styles import Alignment
from pathlib import Path 


def results_table_demographic_group(indicator: str, 
                                    scenario_list: list,
                                    dem_group: list, 
                                    nutrient_label_dict: dict,
                                    results_path: Path,
                                    change_indicator: bool):

  scenario_labels =['Baseline'] + scenario_list
  column_names = [subgroup[0] for subgroup in dem_group]
  if change_indicator:
      column_names = [col + ", change" for col in column_names]
  
  df_results = pd.DataFrame(index =scenario_labels, columns = column_names)
  baseline_results = pd.read_excel(results_path / 'Output/Scenarios/baseline' / f'results_value_baseline.xlsx', index_col=0)
  
  for index, scenario in enumerate(scenario_list):
      if change_indicator:
          plot_data = pd.read_excel(results_path / 'Output/Scenarios' / scenario / f'results_difference_scenario_{index+1}.xlsx', index_col=0)
      else:
          plot_data = pd.read_excel(results_path / 'Output/Scenarios' / scenario / f'results_value_scenario_{index+1}.xlsx', index_col=0)
          
      if 'price' in indicator.lower():
        col_name = 'Price median sim, p'
      else:
        col_name = nutrient_label_dict[f'{indicator}']['title'] + ', ' + nutrient_label_dict[f'{indicator}']['units']
      
      for subgroup in dem_group:
        mean_value = plot_data.loc[subgroup[0], col_name] 
        se_value = plot_data.loc[subgroup[0], f'se_{indicator}']

        lower_value = np.round(mean_value - 1.96*se_value, 2)
        upper_value = np.round(mean_value + 1.96*se_value, 2)

        baseline_value = baseline_results.loc[subgroup[0], col_name]
        baseline_value = np.round(baseline_value, 2)
        
        baseline_se = baseline_results.loc[subgroup[0], f'se_{indicator}']
        baseline_lower_value = np.round(baseline_value - 1.96*baseline_se, 2)
        baseline_upper_value = np.round(baseline_value + 1.96*baseline_se, 2)
       
        if not change_indicator:
          df_results.loc['Baseline', subgroup[0]] = f"{baseline_value}, ({baseline_lower_value}, {baseline_upper_value})"
          df_results.loc[scenario, subgroup[0]] = f"{mean_value}, ({lower_value}, {upper_value})"
        else:
          df_results.loc[scenario, subgroup[0]+", change"] = f"{mean_value}, ({lower_value}, {upper_value})"
          
  return df_results
  
def cumulative_cases_prevented(NCD: str, mortalities: bool, years: int, baseline_path: Path, int_path: Path, filter: tuple, num_seeds: int):

    """
    Computes the number of cases of each disease prevented in all years up to and including the number of years specified by the years parameter.
    
    NCD: either one of "diabetes", "CVD" or "diabetes and CVD"
    mortalities: computes prevented mortalities among those with a disease, as well as all cause mortalities.
    years: number of years over which to compute the prevented cases
    int_path: path to where the scenario simulation health data is saved
    filter: tuple to calculate the impact in a specific demogrphic group
    num_seeds: number of random seeds to use in the uncertainty estimation on the prevented cases.
    
      """

    cases_prevented_total = []
    mortalities_prevented_total = []

    # ensure that the final range is set to 51 so the loop is overall scenarios
    for seed in range(0, num_seeds):
        file_name = f'df_seed_{seed}.parquet'

        baseline_data = pd.read_parquet(baseline_path / f'df_seed_{seed}.parquet')
        int_data = pd.read_parquet(int_path / file_name)

        baseline_data['age'] -= 10
        int_data['age'] -= 10

        dfs = [baseline_data, int_data]
        baseline_data, int_data = filter_dataframes(dfs=dfs, conditions_tuple = filter)

        cases_prevented_total_seed = 0
        mortalities_prevented_total_seed = 0

        for year in range(1, years + 1):
            if mortalities:
                baseline_cases = baseline_data[f'{NCD} mortalities year {year} post'].sum()
                int_cases = int_data[f'{NCD} mortalities year {year} post'].sum()

                baseline_mortalities = baseline_data[f'Total mortalities year {year} post'].sum()
                int_mortalities = int_data[f'Total mortalities year {year} post'].sum()

            else:
                baseline_cases = baseline_data[f'New {NCD} cases year {year}'].sum()
                int_cases = int_data[f'New {NCD} cases year {year}'].sum()

            #baseline_cases_seed.append(baseline_cases)
            cases_prevented_year = baseline_cases - int_cases
            cases_prevented_total_seed += cases_prevented_year

            if mortalities:
              mortalities_prevented_year = baseline_mortalities - int_mortalities
              mortalities_prevented_total_seed +=mortalities_prevented_year

        cases_prevented_total.append(cases_prevented_total_seed)
        mortalities_prevented_total.append(mortalities_prevented_total_seed)

    mean = np.mean(cases_prevented_total)
    lower = np.percentile(cases_prevented_total, 2.5)
    upper = np.percentile(cases_prevented_total, 97.5)

    mean = int(mean)
    lower = int(lower)
    upper = int(upper)

    if mortalities:
      mean_tot = np.mean(mortalities_prevented_total)
      lower_tot = np.percentile(mortalities_prevented_total, 2.5)
      upper_tot = np.percentile(mortalities_prevented_total, 97.5)

      ## Return mortalities among those with each disease and all cause mortality, i.e. if NCD is set to "diabetes" mean wil be mortalities among those with diabetes and mean_tot are all cause mortalities
      return mean, lower, upper, mean_tot, lower_tot, upper_tot

    else:
      return mean, lower, upper

# Returns the prevented all-cause mortalities in each scenario
def all_cause_mortalities(years: int, baseline_path: Path, int_path: Path, filter: tuple, num_seeds: int):

  mortalities_prevented_total = []

  for seed in range(0, num_seeds):
      file_name = f'df_seed_{seed}.parquet'

      baseline_data = pd.read_parquet(baseline_path / f'df_seed_{seed}.parquet')
      int_data = pd.read_parquet(int_path / file_name)

      baseline_data['age'] -= 10
      int_data['age'] -= 10

      dfs = [baseline_data, int_data]
      baseline_data, int_data = filter_dataframes(dfs=dfs, conditions_tuple = filter)

      mortalities_prevented_total_seed = 0

      for year in range(1, years + 1):
        baseline_mortalities = baseline_data[f'Total mortalities year {year} post'].sum()
        int_mortalities = int_data[f'Total mortalities year {year} post'].sum()
        mortalities_prevented_year = baseline_mortalities - int_mortalities
        mortalities_prevented_total_seed +=mortalities_prevented_year

      mortalities_prevented_total.append(mortalities_prevented_total_seed)

  mean = np.mean(mortalities_prevented_total)
  lower = np.percentile(mortalities_prevented_total, 2.5)
  upper= np.percentile(mortalities_prevented_total, 97.5)

  mean = int(mean)
  lower = int(lower)
  upper = int(upper)

  return mean, lower, upper

# Returns the change in BMI in each scenario
def BMI_change(years: int, baseline_path: Path, int_path: Path, filter: tuple, num_seeds: int, df_baseline_nutrients: pd.DataFrame):

    BMI_mean_total = []
    df_sample_weight_total = df_baseline_nutrients.loc[:, 'Sample Weight'].copy()

    # ensure that the final range is set to 51 so the loop is overall scenarios
    for seed in range(0, num_seeds):
        file_name = f'df_seed_{seed}.parquet'

        baseline_data = pd.read_parquet(baseline_path / f'df_seed_{seed}.parquet')
        int_data = pd.read_parquet(int_path / file_name)

        baseline_data['age'] -= 10
        int_data['age'] -= 10

        dfs = [baseline_data, int_data]
        baseline_data, int_data = filter_dataframes(dfs=dfs, conditions_tuple = filter)
        # restrict the sample weights to those in the subgroup
        df_sample_weight = df_sample_weight_total.loc[baseline_data.index]

        # ensure that the change in BMI is calculated relativ to the sample weights at baseline so that they are comparable. 'Sampe Weight' is a dynamical variable in the simualtion so shouldn't be used in the average.
        baseline_BMI = (baseline_data[f'BMI year {years}'] * df_sample_weight).sum() / df_sample_weight.sum()
        int_BMI = (int_data[f'BMI year {years}'] * df_sample_weight).sum() / df_sample_weight.sum()
        change_BMI_seed = int_BMI - baseline_BMI
        BMI_mean_total.append(change_BMI_seed)

    mean = np.mean(BMI_mean_total)
    lower = np.percentile(BMI_mean_total, 2.5)
    upper = np.percentile(BMI_mean_total, 97.5)

    mean = np.round(mean, 2)
    lower = np.round(lower, 2)
    upper = np.round(upper, 2)

    return mean, lower, upper
    

    
def health_outcome_table(indicator: str, dem_group: list, results_path: Path):

  if indicator not in health_outcome_functions:
      raise ValueError(f"Indicator {indicator} is not recognized")

  baseline_path = results_path / 'Output/Percent_reduction/0.0_reduction0.0_dairy_reduction/'

  scenario_path_dict = {"CCC 2030": results_path / 'Output/Percent_reduction/20.0_reduction20.0_dairy_reduction/',
                        'CCC 2050': results_path / 'Output/Percent_reduction/35.0_reduction20.0_dairy_reduction/',
                        'SDG': results_path / 'Output/Max_intake/70.0_max_reduction20.0_dairy_reduction/',
                        'Max red meat 60g/day': results_path /  'Output/Max_intake/60.0_max_reduction20.0_dairy_reduction',
                        'Max red meat 31g/day': results_path / 'Output/Max_intake/31.0_max_reduction20.0_dairy_reduction/',
                        '20% dairy reduction': results_path / 'Output/Percent_reduction/0.0_reduction20.0_dairy_reduction/'
  }

  function = health_outcome_functions[indicator]['function']
  kwargs = health_outcome_functions[indicator]['kwargs']
  base_name = health_outcome_functions[indicator]['column name']

  #columns = [base_name + f", {dem_group}" for dem_group in dem_group_dict['dem_groups'].keys()]
  columns = [base_name + f", {subgroup[0]}" for subgroup in dem_group]
  df_results = pd.DataFrame(index = [f"{i}" for i in scenario_path_dict.keys()], columns = columns)
  

  for scenario, int_path in scenario_path_dict.items():
    for subgroup_condition in dem_group:
        
      column_name = base_name + f", {subgroup_condition[0]}"
      
      mean, lower, upper = function(baseline_path=baseline_path,
                                  int_path=int_path,
                                  filter=subgroup_condition,
                                  **kwargs)

      df_results.loc[f"{scenario}", column_name] = f"{mean}, ({lower}, {upper})"


  return df_results
  
def convert_to_excel(df_dict: dict, file_name: str, save_path: Path):

    file_path = save_path / file_name
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        wb = writer.book
        for indicator, df in df_dict.items():
          if df is not None:
            sheet_name = nutrient_label_dict[indicator]['title']
    
            ws = wb.create_sheet(title=sheet_name)

            # Insert the indicator name in A1
            ws['A1'] = sheet_name
            if indicator in nutrient_label_dict.keys():
              ws["A1"].value += ", "+nutrient_label_dict[indicator]['units']

            ws['A1'].font = Font(bold=True)
            ws['A1'].alignment = Alignment(horizontal='left')

            # Define headers
            dem_groups = [overall_group, age_group, sex_group, simd_group]
            column_index = 2  # Start from column B (Excel index 2)

            for dem_group in dem_groups:
              if column_index == 2:
                absolute_offset = 1
                change_offset = 3
              else:
                absolute_offset = len(dem_group['dem_groups'])+1
                change_offset = len(dem_group['dem_groups'])*2+2

              ws.cell(row=2, column=column_index).value = "Change"
              ws.cell(row=2, column=column_index).font = Font(bold=True)
              ws.cell(row=2, column=column_index).alignment = Alignment(horizontal='center')

              ws.cell(row=2, column=column_index+absolute_offset).value = "Absolute"
              ws.cell(row=2, column=column_index+absolute_offset).font = Font(bold=True)
              ws.cell(row=2, column=column_index+absolute_offset).alignment = Alignment(horizontal='center')

              if column_index>2:
                ws.merge_cells(start_row=2, start_column=column_index, end_row=2, end_column=column_index+len(dem_group["dem_groups"])-1)
                ws.merge_cells(start_row=2, start_column=column_index+absolute_offset, end_row=2, end_column=column_index+absolute_offset+len(dem_group["dem_groups"])-1)

              column_index += change_offset

            # Insert the results table data starting from row 4
            for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), start=3):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx < 4:
                      ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
                      ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)
                    elif c_idx == 1:
                      ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)

                    if c_idx > 1:
                      ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')

            # Adjust the width of the columns
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Get column letter

                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = max_length + 2  # Add some padding
                ws.column_dimensions[col_letter].width = adjusted_width

            # Save the workbook
            #ws.to_excel(writer, sheet_name=sheet_name, index=True)
            wb.save(save_path)

    return
  
  
def results_tables_all_dem_groups(function_mapping: dict, 
                                  demographic_groups: list, 
                                  df_baseline: pd.DataFrame, 
                                  scenario_label_dictionary: dict, 
                                  save_path: Path, 
                                  results_path: Path):

  file_name = 'mSHIFT_SHeS_results'
  file_name = file_name + ".xlsx"
  result_path = os.path.join(save_path, file_name)

  #indicators = [ind for ind in function_mapping.keys()]
  dfs = {indicator: None for indicator in function_mapping.keys()}
  
  # Extract the overall filter for the population
  for dem_filter in demographic_groups:
      if dem_filter['label'] == 'Overall population':
          overall_group = dem_filter.copy()
  
  all_indicators = [ind for ind in function_mapping.keys() if ind not in health_outcome_functions.keys()]
          
  for indicator in health_outcome_functions.keys():
      df_results = health_outcome_table(indicator=indicator, dem_group_dict=overall_group, results_path=results_path)
      df_results[''] = np.nan
      for dem_group_dict in demographic_groups:
        if dem_group_dict != overall_group:
            df_results_dem = health_outcome_table(indicator=indicator, dem_group_dict=dem_group_dict, results_path=results_path)
            df_results_dem[''] = np.nan
            df_results = pd.concat([df_results, df_results_dem], axis=1)
        
      dfs[indicator] = df_results
      convert_to_excel(df_dict = dfs, file_name = file_name, save_path=save_path)
      
  for indicator in all_indicators:
    analysis_function = function_mapping[indicator]

    ##### Compute results for the overall population for the indcator
    results_change = analysis_function(indicator=indicator, 
                                       dem_group_dict=overall_group,
                                       df_baseline=df_baseline,
                                       scenario_label_dictionary=scenario_label_dictionary,
                                       change_indicator=True)
                                       
    results_absolute = analysis_function(indicator=indicator,
                                                    dem_group_dict=overall_group, 
                                                    df_baseline=df_baseline, 
                                                    scenario_label_dictionary=scenario_label_dictionary, 
                                                    change_indicator=False)
                                                    
    results_absolute[''] = np.nan
    results_table = pd.concat([results_change, results_absolute], axis=1)

    ##### Compute results for the demographic groups
    for demographic_group in demographic_groups:
      if demographic_group['label'] != "Overall population":
          results_change_dem_group = analysis_function(indicator=indicator, dem_group_dict=demographic_group, df_baseline=df_baseline, change_indicator=True)
          results_absolute_dem_group = analysis_function(indicator=indicator, dem_group_dict=demographic_group, df_baseline=df_baseline, change_indicator=False)
          results_change_dem_group[''] = np.nan
          results_absolute_dem_group[''] = np.nan
          dem_group_concat = pd.concat([results_change_dem_group, results_absolute_dem_group], axis=1)
    
          dem_group_concat[''] = np.nan
          results_table = pd.concat([results_table, dem_group_concat], axis=1)
    
          dfs[indicator] = results_table
          convert_to_excel(df_dict = dfs, file_name = file_name, save_path=save_path)
    
  return
  
def results_table(
                  dem_group: list, 
                  nutrient_label_dict: dict,
                  error_columns: list,
                  scenario_list: list,
                  scenario_label_map: dict,
                  file_name: str,
                  save_path: Path,
                  results_path: Path
                  ):

  """
  Produces an excel file containing bothe the change in all indicators in all scenarios as well as the absolute new values of each nutrient, environmental impact and cost impact.
  dem_group_dict: dictionary containing the demographic groups to perform the subgroup analysis, one of age_groups, simd_groups, sex_groups or overall_group.
  Saves an excel file with the simulation results for that subgroup
  """

  output_path = save_path
  file_name = file_name + ".xlsx"
  result_path = output_path / file_name

  dfs_health = {} 
  for indicator in health_outcome_functions.keys():
      df_results = health_outcome_table(indicator=indicator, dem_group = dem_group, results_path = results_path)
      dfs_health[indicator] = df_results

  indicators_removed = ['mean', 'acidification', 'biodiversity']
  dfs = {indicator: None for indicator in nutrient_label_dict.keys() if indicator not in error_columns and not any(removal_word in indicator.lower() for removal_word in indicators_removed)}

  for indicator in dfs.keys():

    df_results_change = results_table_demographic_group(
                                                        indicator=indicator,
                                                        scenario_list=scenario_list,
                                                        dem_group = dem_group,
                                                        nutrient_label_dict= nutrient_label_dict,
                                                        change_indicator=True,
                                                        results_path=results_path
                                                        )
                                                        
    df_results_change.rename(index=scenario_label_map, inplace=True)                                                  
    
    df_results = results_table_demographic_group(indicator=indicator, 
                                                 scenario_list=scenario_list,
                                                 dem_group = dem_group, 
                                                 nutrient_label_dict= nutrient_label_dict,
                                                 change_indicator=False,
                                                 results_path=results_path
                                                 )
                                                 
    df_results.rename(index=scenario_label_map, inplace=True)  
    
    df_results_change[''] = np.nan
    df_results_combined = pd.concat([df_results_change, df_results], axis=1)
    dfs[indicator] = df_results_combined
    
  dfs = dfs | dfs_health
  
  env_col_rename = {'median_price_sim': "Cost",
                    'median_Eut': "Eutrophication",
                    'median_GHG': "Greenhouse gas emissions",
                    'median_Land': "Land use",
                    'median_WaterUse': "Freshwater use"}

  with pd.ExcelWriter(result_path, engine='openpyxl') as writer:
    wb = writer.book
    for indicator, data_frame in dfs.items():
        # Determine the sheet name
        if indicator in nutrient_label_dict.keys() and indicator not in env_col_rename.keys():
            sheet_name = nutrient_label_dict[indicator]['title'] + ", " + nutrient_label_dict[indicator]['units']
        elif sheet_name in env_col_rename.keys():
            sheet_name = env_col_rename[indicator] + ", " + nutrient_label_dict[indicator]['units']
            if 'eutrophication' in sheet_name.lower():
                sheet_name = sheet_name.replace("gPO$_3$e", "gPO4e")
            elif 'greenhouse' in sheet_name.lower():
                sheet_name = sheet_name.replace('kgCO$_2$e', 'kgCO2e')
            else:
                pass
            
                
        else:
            sheet_name = indicator
            
        sheet_name = sheet_name.replace('$\\mu$', 'u')

        # Write the DataFrame to the Excel sheet
        data_frame.to_excel(writer, sheet_name=sheet_name, index=True)
        worksheet = writer.sheets[sheet_name]

        # Adjust the column widths 
        for column_cells in worksheet.columns:
            max_length = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            worksheet.column_dimensions[col_letter].width = max_length + 2



  return


# For submitted results num_seeds=51 and years=10
num_seeds = 51
years = 10

df_baseline = pd.read_parquet('data/df_baseline.parquet')

health_outcome_functions = {'BMI': {'function': BMI_change,
                                    'column name': 'Per capita change in BMI after 10 years (95 percent UI)',
                                    'kwargs': {'years': years,
                                                'num_seeds': num_seeds,
                                                'df_baseline_nutrients': df_baseline
                                    }
                                    },
    
    
    
                            'CVD':  {'function': cumulative_cases_prevented,
                                                             'column name': 'Prevented CVD cases after 10 years (95 percent UI)',
                                                             'kwargs' : {'NCD': "CVD",
                                                                         'mortalities': False,
                                                                         'years': years,
                                                                            'num_seeds': num_seeds}
                                                              },

                          'diabetes': {'function': cumulative_cases_prevented,
                                       'column name': 'Prevented diabetes cases after 10 years (95 percent UI)',
                                       'kwargs': {'NCD': "diabetes",
                                                   'mortalities': False,
                                                    'years': years,
                                                    'num_seeds': num_seeds
                                       }
                                         },
                          'all cause mortality': {'function': all_cause_mortalities,
                                                  'column name': "Prevented all cause mortality after 10 years (95 percent UI)",
                                                  'kwargs': {'years': years,
                                                            'num_seeds': num_seeds}
                                                  }

                            

                            }